"""
build_spotcheck.py — Human Spot-Check for §A.7 (FIXED)

Run in your sdt_calibration venv:
    python build_spotcheck.py

Reads:
    data/triviaqa_5000.json, data/nq_3000.json
    results/paradigm_a/*.jsonl (T=1.0 trials for all 3 models x 2 datasets)

Outputs:
    spotcheck.xlsx — ready for manual scoring
"""

import json
import random
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BASE_DIR = Path(r"C:\sdt_calibration")
SEED = 42
N_TRIVIAQA = 250
N_NQ = 150
MODELS = ["llama3_instruct", "mistral_instruct", "llama3_base"]
MODEL_SHORT = {"llama3_instruct": "L3-Inst", "mistral_instruct": "Mistral", "llama3_base": "L3-Base"}

HEADER_FILL = PatternFill("solid", fgColor="4472C4")
HEADER_FONT = Font(bold=True, size=10, name="Arial", color="FFFFFF")
BODY_FONT = Font(size=10, name="Arial")
WRAP = Alignment(wrap_text=True, vertical="top")
CENTRE = Alignment(horizontal="center", vertical="center")
THIN = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
GREEN = PatternFill("solid", fgColor="C6EFCE")
RED = PatternFill("solid", fgColor="FFC7CE")
YELLOW = PatternFill("solid", fgColor="FFEB9C")
LIGHT_GREY = PatternFill("solid", fgColor="F2F2F2")


def load_dataset(name: str) -> list:
    path = BASE_DIR / "data" / f"{'triviaqa_5000' if name == 'triviaqa' else 'nq_3000'}.json"
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)


def get_aliases_display(item: dict, dataset: str) -> str:
    """Format ground truth aliases for display.

    Data format (from prepare_datasets.py — flattened, NOT nested):
      TriviaQA: answer_value, answer_aliases, answer_normalized_aliases
      NQ: answer_value, answer_aliases
    """
    if dataset == "triviaqa":
        aliases = list(item.get("answer_aliases", []))
        value = item.get("answer_value", "")
        if value and value not in aliases:
            aliases = [value] + aliases
        return " | ".join(aliases[:8]) if aliases else str(value)
    else:  # nq
        aliases = list(item.get("answer_aliases", []))
        value = item.get("answer_value", "")
        if value and value not in aliases:
            aliases = [value] + aliases
        return " | ".join(aliases[:8]) if aliases else str(value)


def load_t10_trials(model: str, dataset: str) -> dict:
    path = BASE_DIR / "results" / "paradigm_a" / f"{model}_{dataset}.jsonl"
    results = {}
    with open(path, "r", encoding="utf-8") as f:
        for line in f:
            trial = json.loads(line)
            if abs(trial["temperature"] - 1.0) < 0.01:
                results[trial["question_index"]] = trial
    return results


def main():
    triviaqa = load_dataset("triviaqa")
    nq = load_dataset("nq")
    print(f"Loaded TriviaQA ({len(triviaqa)}), NQ ({len(nq)})")

    # Verify data format
    sample = triviaqa[0]
    assert "answer_aliases" in sample or "answer_value" in sample, \
        f"Unexpected data format. Keys: {list(sample.keys())}"
    print(f"TriviaQA keys: {list(sample.keys())}")
    print(f"  answer_value: {sample.get('answer_value', 'MISSING')}")
    aliases_sample = sample.get('answer_aliases', [])
    print(f"  answer_aliases ({len(aliases_sample)}): {aliases_sample[:3]}...")

    random.seed(SEED)
    tqa_idx = sorted(random.sample(range(len(triviaqa)), N_TRIVIAQA))
    nq_idx = sorted(random.sample(range(len(nq)), N_NQ))
    print(f"Sampled {len(tqa_idx)} TriviaQA + {len(nq_idx)} NQ = {len(tqa_idx) + len(nq_idx)}")

    all_data = {}
    for model in MODELS:
        for ds in ["triviaqa", "nq"]:
            key = f"{model}_{ds}"
            all_data[key] = load_t10_trials(model, ds)
            print(f"  {key}: {len(all_data[key])} T=1.0 trials")

    wb = Workbook()
    ws = wb.active
    ws.title = "Spot Check"

    fixed = ["#", "Dataset", "Q_Idx", "Question", "Ground_Truth"]
    per_model = ["Answer", "Stripped", "Auto", "Match", "Sim", "Alias", "Ref"]
    scoring = ["Human", "Disagree?"]
    final = ["Notes"]

    headers = list(fixed)
    for m in MODELS:
        s = MODEL_SHORT[m]
        for cn in per_model:
            headers.append(f"{s}_{cn}")
        for cn in scoring:
            headers.append(f"{s}_{cn}")
    headers.extend(final)

    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", wrap_text=True, vertical="center")
        cell.border = THIN
    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "F2"

    samples = [(idx, "triviaqa", triviaqa) for idx in tqa_idx] + \
              [(idx, "nq", nq) for idx in nq_idx]

    gt_populated = 0
    for row_i, (q_idx, ds, dataset) in enumerate(samples, 2):
        item = dataset[q_idx]
        aliases_str = get_aliases_display(item, ds)
        if aliases_str:
            gt_populated += 1

        ws.cell(row=row_i, column=1, value=row_i - 1).font = BODY_FONT
        ws.cell(row=row_i, column=2, value="TQA" if ds == "triviaqa" else "NQ").font = BODY_FONT
        ws.cell(row=row_i, column=3, value=q_idx).font = BODY_FONT

        q_cell = ws.cell(row=row_i, column=4, value=item["question"])
        q_cell.font = BODY_FONT
        q_cell.alignment = WRAP

        gt_cell = ws.cell(row=row_i, column=5, value=aliases_str)
        gt_cell.font = BODY_FONT
        gt_cell.alignment = WRAP

        col = 6
        for model in MODELS:
            key = f"{model}_{ds}"
            trial = all_data[key].get(q_idx)

            if trial:
                ws.cell(row=row_i, column=col, value=trial.get("generated_text", "")).alignment = WRAP
                ws.cell(row=row_i, column=col + 1, value=trial.get("stripped_text", "")).alignment = WRAP

                score = trial.get("correct", False)
                sc = ws.cell(row=row_i, column=col + 2, value=int(score))
                sc.fill = GREEN if score else RED
                sc.alignment = CENTRE

                ws.cell(row=row_i, column=col + 3, value=trial.get("match_type", ""))

                sim = trial.get("best_similarity")
                if sim is not None:
                    ws.cell(row=row_i, column=col + 4, value=round(sim, 3))

                ws.cell(row=row_i, column=col + 5, value=trial.get("matched_alias", "")).alignment = WRAP
                ws.cell(row=row_i, column=col + 6, value=int(trial.get("refusal_flag", False)))

                hc = ws.cell(row=row_i, column=col + 7)
                hc.fill = YELLOW
                hc.alignment = CENTRE

                ws.cell(row=row_i, column=col + 8).alignment = CENTRE
            else:
                ws.cell(row=row_i, column=col, value="[MISSING]")

            for c in range(col, col + 9):
                ws.cell(row=row_i, column=c).border = THIN
                ws.cell(row=row_i, column=c).font = BODY_FONT

            col += 9

        ws.cell(row=row_i, column=col).border = THIN

        for c in range(1, 6):
            ws.cell(row=row_i, column=c).border = THIN

        if row_i % 2 == 0:
            for c in range(1, 6):
                ws.cell(row=row_i, column=c).fill = LIGHT_GREY

    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 6
    ws.column_dimensions["C"].width = 7
    ws.column_dimensions["D"].width = 45
    ws.column_dimensions["E"].width = 35

    col = 6
    for _ in MODELS:
        ws.column_dimensions[get_column_letter(col)].width = 22
        ws.column_dimensions[get_column_letter(col + 1)].width = 22
        ws.column_dimensions[get_column_letter(col + 2)].width = 6
        ws.column_dimensions[get_column_letter(col + 3)].width = 8
        ws.column_dimensions[get_column_letter(col + 4)].width = 6
        ws.column_dimensions[get_column_letter(col + 5)].width = 18
        ws.column_dimensions[get_column_letter(col + 6)].width = 5
        ws.column_dimensions[get_column_letter(col + 7)].width = 7
        ws.column_dimensions[get_column_letter(col + 8)].width = 9
        col += 9
    ws.column_dimensions[get_column_letter(col)].width = 30

    # Instructions sheet
    ws2 = wb.create_sheet("Instructions")
    instructions = [
        ("HUMAN SPOT-CHECK — A.7", Font(bold=True, size=14, name="Arial")),
        ("", None),
        ("Purpose: Verify automated scoring pipeline accuracy.", Font(size=11, name="Arial")),
        ("", None),
        ("PROCEDURE:", Font(bold=True, size=11, name="Arial")),
        ("1. For each row, examine Question + Ground_Truth.", None),
        ("2. For each model, compare Stripped answer against Ground_Truth aliases.", None),
        ("3. Fill Human column: 1 = correct answer, 0 = incorrect.", None),
        ("4. If Human != Auto, put YES in the Disagree? column.", None),
        ("5. Add Notes for ambiguous cases.", None),
        ("", None),
        ("SCORING RULES:", Font(bold=True, size=11, name="Arial")),
        ("- Accept reasonable variations: 'Mt. Everest' = 'Mount Everest'", None),
        ("- Accept minor typos if the answer is clearly intended", None),
        ("- Reject wrong answers even if superficially similar", None),
        ("- Reject refusals (Auto should already score these as 0)", None),
        ("- If model gives additional correct info beyond the answer, score correct", None),
        ("", None),
        ("METRICS (compute after scoring):", Font(bold=True, size=11, name="Arial")),
        ("  False-match rate = (Auto=1 & Human=0) / N(Auto=1)  [too generous]", None),
        ("  Missed-match rate = (Auto=0 & Human=1) / N(Auto=0)  [too strict]", None),
        ("  Threshold: either rate > 3% triggers pipeline revision", None),
        ("", None),
        (f"Sample: {N_TRIVIAQA} TriviaQA + {N_NQ} NQ = {N_TRIVIAQA + N_NQ} questions", None),
        (f"Models: 3 (all at T=1.0). Total judgments: {(N_TRIVIAQA + N_NQ) * 3}", None),
        (f"Seed: {SEED}", None),
    ]
    for i, (text, font) in enumerate(instructions, 1):
        cell = ws2.cell(row=i, column=1, value=text)
        cell.font = font if font else Font(size=11, name="Arial")
    ws2.column_dimensions["A"].width = 80

    # Summary sheet
    ws3 = wb.create_sheet("Summary")
    ws3.cell(row=1, column=1, value="Spot-Check Summary").font = Font(bold=True, size=14, name="Arial")
    ws3.cell(row=3, column=1, value="Fill this in after scoring:").font = Font(bold=True, size=11, name="Arial")

    summary_rows = [
        (5, "Metric", "L3-Inst", "Mistral", "L3-Base", "Overall"),
        (6, "N scored", "", "", "", ""),
        (7, "N Auto=1 & Human=1 (true pos)", "", "", "", ""),
        (8, "N Auto=1 & Human=0 (false match)", "", "", "", ""),
        (9, "N Auto=0 & Human=1 (missed match)", "", "", "", ""),
        (10, "N Auto=0 & Human=0 (true neg)", "", "", "", ""),
        (11, "False-match rate", "", "", "", ""),
        (12, "Missed-match rate", "", "", "", ""),
        (13, "Overall agreement %", "", "", "", ""),
    ]
    for row_data in summary_rows:
        rn = row_data[0]
        for ci, val in enumerate(row_data[1:], 1):
            cell = ws3.cell(row=rn, column=ci, value=val)
            cell.border = THIN
            if rn == 5:
                cell.font = HEADER_FONT
                cell.fill = HEADER_FILL
            else:
                cell.font = Font(size=10, name="Arial")
                if ci > 1:
                    cell.fill = YELLOW
    ws3.column_dimensions["A"].width = 35
    for c in ["B", "C", "D", "E"]:
        ws3.column_dimensions[c].width = 14

    out_path = BASE_DIR / "spotcheck.xlsx"
    wb.save(out_path)
    print(f"\nGround_Truth populated: {gt_populated}/{len(samples)}")
    print(f"Saved: {out_path}")
    print(f"Rows: {len(samples)}")
    print(f"Judgments needed: {len(samples) * len(MODELS)}")


if __name__ == "__main__":
    main()
